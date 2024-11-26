from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from .forms import UserLoginForm, UserRegistrationForm, UserProfileForm, DateForm, ProductWeightForm
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from .models import Diary, ProductsDiaries, Product

from .utils.calcs import calculate_total_bju, calculate_each_product_bju, calc_rsk
from datetime import date, datetime
from babel.dates import format_date

import openpyxl
from openpyxl.styles import Font, Alignment

@login_required
def export_diaries_to_xlsx(request):
    # Создаем Excel-книгу и активный лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дневники"
    
    # Устанавливаем шапку
    headers = ["Дата", "Калории (ккал)", "Жиры (г)", "Белки (г)", "Углеводы (г)"]
    ws.append(headers)
    
    # Формат заголовка
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    
    # Получаем дневники текущего пользователя
    diaries = Diary.objects.filter(user=request.user).order_by("date")
    
    row = 2  # Начальная строка для данных
    
    for diary in diaries:
        # Считаем тотальные показатели для текущей даты
        products = diary.product_entries.all()
        total_calories = sum(entry.product.calories * entry.weight / 100 for entry in products)
        total_fats = sum(entry.product.fats * entry.weight / 100 for entry in products)
        total_proteins = sum(entry.product.prots * entry.weight / 100 for entry in products)
        total_carbos = sum(entry.product.carbos * entry.weight / 100 for entry in products)
        
        # Добавляем строку с общей информацией по дате
        ws.append([
            diary.date.strftime("%d.%m.%Y"),  # Дата в формате ДД.ММ.ГГГГ
            round(total_calories, 2),
            round(total_fats, 2),
            round(total_proteins, 2),
            round(total_carbos, 2)
        ])
        
        # Форматируем строку с общей информацией
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        
        row += 1  # Переход на следующую строку
        
        # Добавляем данные по каждому продукту
        for entry in products:
            ws.append([
                entry.product.name,
                round(entry.product.calories * entry.weight / 100, 2),
                round(entry.product.fats * entry.weight / 100, 2),
                round(entry.product.prots * entry.weight / 100, 2),
                round(entry.product.carbos * entry.weight / 100, 2)
            ])
            row += 1
        
        
    
    # Сохраняем файл и возвращаем его пользователю
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=diaries_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    wb.save(response)
    return response



@login_required
def index(request):
    products = None
    message = None
    enriched_products = None
    total_bju = None
    is_admin = None

    # Админ ли пользователь
    if request.user.is_superuser:
        is_admin = True

    if request.method == 'POST':
        form = DateForm(request.POST)
        if form.is_valid():
            date_selected = form.cleaned_data['date']
            # обновление параметра date после каждого переключения
            return redirect(f"{reverse('diary:index')}?date={date_selected}")
    else:
        date_selected = request.GET.get('date')
        print(date_selected)
        if date_selected:
            try:
                date_selected = date.fromisoformat(date_selected)
            except ValueError:
                date_selected = datetime.strptime(date_selected, "%b. %d, %Y").date()
        else:
            date_selected = date.today() 
        form = DateForm(initial={'date': date_selected})

    # Проверяем, есть ли дневник на указанную дату
    try:
        diary = Diary.objects.get(user=request.user, date=date_selected)

        if not diary.product_entries.exists():
            diary.delete()
            print(f"Удален пустой дневник за {diary.date}")
        else: 
            products = diary.product_entries.all()  # Получаем все записи продуктов из дневника

            # расчет bju для каждого продукта с учетом веса
            enriched_products = calculate_each_product_bju(products)

            calc_rsk(enriched_products, request.user.rsk)

            # расчет total bju
            total_bju = calculate_total_bju(enriched_products, request.user.calorie_goal)

    except Diary.DoesNotExist:
        message = "Дневник пуст! Добавьте продукты кнопкой +"

    # Отображение даты
    if date_selected == date.today():
        date_display = 'Сегодня'
    else:
        date_display = format_date(date_selected, format="d MMMM yyyy", locale="ru")

    context = {
        'form': form,
        'products': enriched_products,
        'message': message,
        'total_bju': total_bju,
        'date': date_display,
        'is_admin': is_admin,
    }
    return render(request, 'diary/diary.html', context)



def login(request):
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user:
                auth_login(request, user)
                # Получаем значение 'next' из параметра запроса
                next_url = request.GET.get('next')
                if next_url:  # Если параметр 'next' существует
                    return redirect(next_url)
                else:  # Если параметра 'next' нет, перенаправляем на главную страницу дневника
                    return redirect('diary:index')
    else:
        form = UserLoginForm()

    context = {'form': form}
    return render(request, 'diary/login.html', context)


def registration(request):
    if request.method == 'POST':
        form = UserRegistrationForm(data=request.POST)
        if form.is_valid():
            form.save()
            return redirect('diary:login')
    else:
        form = UserRegistrationForm()
    context = {'form': form}

    return render(request, 'diary/registration.html', context)

@login_required
def profile(request):
    if request.method == 'POST':
        profile_form = UserProfileForm(request.POST, instance=request.user)
        if profile_form.is_valid():
            profile_form.save()
            return redirect('diary:profile')  
    else:
        profile_form = UserProfileForm(instance=request.user)

    # Счётчик дней, где достигнута или превышена цель калорий
    calorie_goal_days = 0

    for diary in request.user.diaries.all():
        total_calories = sum(
            entry.product.calories * entry.weight / 100 for entry in diary.product_entries.all()
        )
        if total_calories >= request.user.calorie_goal:
            calorie_goal_days += 1

    # Счетчик средних калорий в день
    total_calories_by_day = 0
    for diary in request.user.diaries.all():
        total_calories_by_day += sum(
            entry.product.calories * entry.weight / 100 for entry in diary.product_entries.all()
        )

    try:
        avg_calories = total_calories_by_day / request.user.diaries.count()
    except ZeroDivisionError:
        avg_calories = 0

    context = {
        'profile_form': profile_form,
        'username': request.user.username,
        'email': request.user.email,
        'age': request.user.age,
        'rsk': request.user.rsk,
        'days_filled': request.user.diaries.count(),
        'goal_riched_days': calorie_goal_days,
        'avg_calories': avg_calories,
    }
    return render(request, 'diary/profile.html', context)


def logout(request):
    auth_logout(request)
    return HttpResponseRedirect(reverse('diary:index'))


@login_required
def product_edit(request, entry_id):
    entry = get_object_or_404(ProductsDiaries, id=entry_id)

    if request.method == 'POST':
        if 'delete' in request.POST:
            entry.delete()
            return redirect(f"{reverse('diary:index')}?date={entry.diary.date}")
        elif 'save' in request.POST:
            entry.weight = request.POST.get('weight')
            entry.save()
            return redirect(f"{reverse('diary:index')}?date={entry.diary.date}")

    if entry.diary.date == date.today():
        display_date = 'Сегодня'
    else:
        display_date = format_date(entry.diary.date, format="d MMMM yyyy", locale="ru")

    context = {'entry': entry,
               'date': display_date,
               }
    return render(request, 'diary/edit.html', context)


@login_required
def add_product(request):
    date_selected = request.GET.get('date')
    not_found_msg = None
    
    if not date_selected:
        date_selected = date.today()
    else:
        try:
            date_selected = datetime.strptime(date_selected, "%b. %d, %Y").date()
        except ValueError:
            print("НЕКОРЕКТНАЯ ДАТА: ", date_selected)
            return redirect('diary:index')  # Вернуться на страницу дневника, если дата некорректна

    diary, created = Diary.objects.get_or_create(user=request.user, date=date_selected)

    if 'search' in request.GET:
        query = request.GET.get('search')
        print(f"Поисковый запрос: {query}")
        products = Product.objects.filter(name__icontains=query)[:15]

        if products.count() == 0:
            not_found_msg = 'Ничего не найдено'        
    else:
        products = Product.objects.all()[:15]

    # Отображение даты
    if date_selected == date.today():
        date_display = 'Сегодня'
    else:
        date_display = format_date(date_selected, format="d MMMM yyyy", locale="ru")
   
    context = {
        'date': date_selected,
        'products': products,
        'date_display': date_display,
        'msg': not_found_msg,
    }
    return render(request, 'diary/add_product.html', context)


@login_required
def product_card(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    date_selected = request.GET.get('date')
    date_selected = datetime.strptime(date_selected, "%b. %d, %Y").date()

    if request.method == 'POST':
        diary = request.user.diaries.get(date=date_selected)
        form = ProductWeightForm(request.POST)
        if form.is_valid():
            weight = form.cleaned_data['weight']
            ProductsDiaries.objects.create(
                diary = diary,
                product=product,
                weight=weight
            )

            return redirect(f"{reverse('diary:index')}?date={date_selected}")
    else:
        form = ProductWeightForm()

    # Отображение даты
    if date_selected == date.today():
        date_display = 'Сегодня'
    else:
        date_display = format_date(date_selected, format="d MMMM yyyy", locale="ru")

    context = {
        'product': product,
        'date': date_selected,
        'form': form,
        'date_display': date_display,
    }

    return render(request, 'diary/product_card.html', context)